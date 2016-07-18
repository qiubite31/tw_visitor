# coding=UTF-8
import os
import sqlite3
from openpyxl import load_workbook
from xlrd import open_workbook
from xlrd.sheet import ctype_text   
import numpy as np
import pandas as pd

#conn = sqlite3.connect('visitor.db')

print(os.path.dirname(__file__))
wb = open_workbook('C:/Users/Dragon/tour_statistic/mysite/tours/vistor_history/1602.xls', formatting_info=True)

sheet = wb.sheet_by_name('Sheet3')

#print(type(sheet.cell_value(0,0)))
print(sheet.cell_value(1,3))

#print(sheet.row_values(5))

# Read Column Title
column_title = []
for i in range(0, sheet.ncols):
    cell_value = sheet.cell(1, i).value
    if not (cell_value == '' or 'Residence' in cell_value or 'Total' in cell_value):
        column_title.append(cell_value.replace('\n', ' '))

print(column_title)

index_title = []
for i in range(0, sheet.nrows):
    cell_value = sheet.cell(i, 1).value
    if 'Total' in cell_value or 'Total' in sheet.cell_value(i, 2):
        continue
    if '東南亞地區' in cell_value or cell_value == '':
        cell_value_shift = sheet.cell_value(i, 2)
        if not cell_value_shift == '':
            index_title.append(cell_value_shift)
    else:
        index_title.append(cell_value)
print(index_title)

data_list = []
for i in range(0, sheet.ncols-1):
    data = []
    for j in range(0, sheet.nrows-1):
        if sheet.cell(j, i).ctype == 2:
            if not('Total' in sheet.cell_value(j, 2) or 'Total' in sheet.cell_value(j, 1)):
                data.append(sheet.cell(j, i).value)
            #print(sheet.cell(j, i).value + ' ' + str(sheet.cell(j, i).ctype))
        if j == sheet.nrows - 2 and len(data) > 0:
            data_list.append(data)

#print(data_list)

visitor_data = pd.DataFrame(data_list, column_title, index_title)
#print(visitor_data.index)
print(visitor_data.head)

#print(sheet.ncols)
#print(sheet.nrows)

"""
row = sheet.row(1)  # 1st row 
# Print 1st row values and types
print('(Column #) type:value')
for idx, cell_obj in enumerate(row):
    cell_type_str = ctype_text.get(cell_obj.ctype, 'unknown type')
    print('(%s) %s %s' % (idx, cell_type_str.encode('utf-8'), cell_obj.value))
"""